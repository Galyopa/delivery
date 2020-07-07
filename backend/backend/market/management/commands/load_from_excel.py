from django.core.management.base import BaseCommand, CommandError
from backend.settings import DATA_DIR
from openpyxl import load_workbook
from market.models import Category, Product


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print('Start importing from excel %s' % DATA_DIR)
        # open the book
        wb = load_workbook(DATA_DIR + '/price.xlsx')
        # extract the first sheet
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        category = None

        for count in range(1, sheet.max_row + 1):
            item = sheet.cell(row=count, column=3).value
            id = sheet.cell(row=count, column=2).value
            if id == None:
                print('Create a new category')
                category = Category()
                category.name = item
                category.save()
            else:
                print('Create a new good')
                if category:
                    product = Product()
                    product.name = item
                    product.category = category
                    product.save()
